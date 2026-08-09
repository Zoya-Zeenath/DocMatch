import streamlit as st
import pdfplumber
import pandas as pd
import pytesseract
from pdf2image import convert_from_bytes
from PIL import Image, ImageEnhance
import io
import re
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

st.set_page_config(page_title="DocMatch", layout="wide")
st.title("DocMatch")
st.caption("Resume & Aadhaar Extractor")

# ──────────────────────────────────────────────
# OCR HELPERS
# ──────────────────────────────────────────────

def pdf_to_text(file_bytes):
    text = ""
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    except Exception:
        pass
    if len(text.strip()) < 30:
        try:
            for img in convert_from_bytes(file_bytes, dpi=300):
                gray = ImageEnhance.Contrast(img.convert("L")).enhance(2.0)
                text += pytesseract.image_to_string(gray) + "\n"
        except Exception:
            pass
    return text


def aadhaar_multipass_ocr(file_bytes):
    """
    5 OCR passes with different crop/contrast settings.
    Returns (primary_text, [all_texts]).
    Tested on real Aadhaar photo: correctly reads 'Arabiya Sultana'.
    """
    img = Image.open(io.BytesIO(file_bytes))
    w, h = img.size
    results = []
    for crop_r, contrast in [(0.60, 2.5), (0.60, 3.0), (0.62, 2.5), (0.58, 2.5), (0.60, 2.0)]:
        crop = img.crop((0, 0, w, int(h * crop_r)))
        up   = crop.resize((w * 3, int(h * crop_r) * 3), Image.LANCZOS)
        gray = ImageEnhance.Contrast(up.convert("L")).enhance(contrast)
        results.append(pytesseract.image_to_string(gray, config="--oem 3 --psm 6"))
    return results[0], results


# ──────────────────────────────────────────────
# RESUME EXTRACTORS
# ──────────────────────────────────────────────

def get_resume_name(text):
    skip = {"resume", "cv", "curriculum vitae", "profile", "summary",
            "objective", "contact", "email", "phone", "mobile", "address", "name"}
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines[:12]:
        if any(k in line.lower() for k in skip):
            continue
        if re.search(r"[^a-zA-Z\s\.\-]", line):
            continue
        if 2 <= len(line.split()) <= 5 and re.match(r"^[A-Z][a-z]+(\s[A-Z][a-z]+)+$", line):
            return line
    for line in lines[:8]:
        if 2 <= len(line.split()) <= 5 and re.match(r"^[A-Za-z\s\.\-]+$", line):
            return line.title()
    return ""

def get_resume_email(text):
    m = re.search(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return m.group(0) if m else ""

def get_resume_phone(text):
    """
    Extracts Indian mobile numbers in any format, e.g.:
      +91-9876543210  |  +91 9876543210  |  +919876543210
      0091-9876543210 |  09876543210
      9876543210      |  98765 43210
      9876-543210     |  98765-43210
      (98765) 43210
    Always returns clean 10 digits starting with 6-9.
    """
    def clean(raw):
        digits = re.sub(r"[^\d]", "", raw)
        if len(digits) == 12 and digits.startswith("91"):
            digits = digits[2:]
        elif len(digits) == 13 and digits.startswith("091"):
            digits = digits[3:]
        elif len(digits) == 14 and digits.startswith("0091"):
            digits = digits[4:]
        elif len(digits) == 11 and digits.startswith("0"):
            digits = digits[1:]
        if len(digits) == 10 and digits[0] in "6789":
            return digits
        return ""

    patterns = [
        # Labelled with any prefix format
        r"(?:phone|mobile|mob|cell|ph|tel|contact|whatsapp)[^\d]{0,5}(\+?0{0,2}91[\s\-\.]?[6-9][\d\s\-\.]{9,14})",
        r"(?:phone|mobile|mob|cell|ph|tel|contact|whatsapp)[^\d]{0,5}(0?[6-9][\d\s\-\.]{9,13})",
        # +91 or 0091
        r"(\+91[\s\-\.]?[6-9][\d\s\-\.]{9,12})",
        r"(0091[\s\-\.]?[6-9][\d\s\-\.]{9,12})",
        # Leading trunk 0: 09876543210
        r"\b(0[6-9]\d{9})\b",
        # Spaced or hyphenated: 98765 43210 / 9876-543210 / 98765-43210
        r"\b([6-9]\d{4}[\s\-\.]\d{5})\b",
        r"\b([6-9]\d{3}[\s\-\.]\d{6})\b",
        r"\b([6-9]\d{2}[\s\-\.]\d{7})\b",
        # Bracketed: (98765) 43210
        r"\(([6-9]\d{4})\)[\s\-\.]?(\d{5})",
        # Plain 10-digit
        r"\b([6-9]\d{9})\b",
    ]

    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            raw = "".join(g for g in m.groups() if g)
            result = clean(raw)
            if result:
                return result
    return ""

def get_resume_college(text):
    keywords = ["university", "college", "institute", "institution", "iit", "nit",
                "polytechnic", "engineering college", "academy", "school of",
                "b.tech", "b.e.", "bachelor", "master", "m.tech", "mba"]

    def clean_college(line):
        # Remove leading bullets / special chars (•, *, -, –, etc.)
        line = re.sub(r"^[\s\•\*\-\–\—\►\▪\·]+", "", line)
        # Remove date patterns like "Jan2021-May2024", "March 2025", "2020-2024"
        line = re.sub(r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s\-]?\d{4}\b", "", line, flags=re.IGNORECASE)
        line = re.sub(r"\b\d{4}\s*[-–]\s*\d{4}\b", "", line)
        line = re.sub(r"\b\d{4}\b", "", line)
        # Remove board/result suffixes like "APOSS Board", "– March 2025"
        line = re.sub(r"[,\-–]\s*(aposs|cbse|icse|state board|board|result|march|april|may|june|july)[^,]*", "", line, flags=re.IGNORECASE)
        # Remove class/grade indicators like "(Class XII)", "Class 12"
        line = re.sub(r"\(?\bclass\s+(xii|xi|x|12|11|10)\)?", "", line, flags=re.IGNORECASE)
        # Remove degree prefixes like "Intermediate", "B.Tech –", "B.E. -"
        line = re.sub(r"^(intermediate|b\.?tech|b\.?e\.?|m\.?tech|m\.?e\.?|bsc|msc|mba|bca|mca|diploma)\s*[\-–,]?\s*", "", line, flags=re.IGNORECASE)
        # Remove anything inside parentheses
        line = re.sub(r"\(.*?\)", "", line)
        # Remove leftover special characters except letters, spaces, commas, dots, &
        line = re.sub(r"[^a-zA-Z\s,\.&]", "", line)
        # Collapse multiple spaces
        line = re.sub(r"\s{2,}", " ", line).strip().strip(",").strip()
        return line

    for line in text.splitlines():
        if any(kw in line.lower() for kw in ["university", "college", "institute"]):
            cleaned = clean_college(line)
            if 3 < len(cleaned) < 120:
                return cleaned

    for line in text.splitlines():
        if any(kw in line.lower() for kw in keywords):
            cleaned = clean_college(line)
            if 3 < len(cleaned) < 120:
                return cleaned

    return ""


# ──────────────────────────────────────────────
# AADHAAR EXTRACTORS
# Layout: To → <Regional script> → <English Name> → D/O → #house → street → PIN
#                                                                           → phone
# ──────────────────────────────────────────────

def _is_name(s):
    s = s.strip()
    if len(s) < 4: return False
    words = s.split()
    if not (2 <= len(words) <= 6): return False
    return all(re.match(r"^[A-Za-z\.\-]+$", w) for w in words)

def _parse_name_single(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    for i, line in enumerate(lines):
        if re.fullmatch(r"To\.?", line, re.IGNORECASE):
            cands = []
            cand_idx = []
            for j in range(i + 1, min(i + 9, len(lines))):
                c = re.sub(r"^[\-\—\s]+|[\-\—\s]+$", "", lines[j]).strip()
                if re.match(r"^(D/O|S/O|W/O|C/O|O:|#\d)", c, re.IGNORECASE):
                    break
                if _is_name(c):
                    cands.append(c.title())
                    cand_idx.append(j)

            if len(cands) >= 2:
                # 1st = regional script rendered as latin, 2nd = English name
                return cands[1]

            if len(cands) == 1:
                # Only one candidate — check if D/O comes right after it.
                # If yes: the regional line wasn't detected as a name, so this IS the English name.
                nxt = cand_idx[0] + 1
                if nxt < len(lines) and re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", lines[nxt], re.IGNORECASE):
                    return cands[0]
                # Otherwise skip — it's probably the regional script garbage line

    # Fallback: line just before D/O
    for i, line in enumerate(lines):
        if re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", line, re.IGNORECASE):
            for j in range(i - 1, max(i - 4, -1), -1):
                c = re.sub(r"^[\-\—\s]+|[\-\—\s]+$", "", lines[j]).strip()
                if _is_name(c): return c.title()

    # Fallback: before DOB
    for i, line in enumerate(lines):
        if re.search(r"\bDOB\b|\bDate of Birth\b|\bD\.O\.B\b", line, re.IGNORECASE):
            for j in range(i - 1, max(i - 4, -1), -1):
                c = re.sub(r"^[\-\—\s]+|[\-\—\s]+$", "", lines[j]).strip()
                if _is_name(c): return c.title()
    return ""

def get_aadhaar_name(all_texts):
    """Vote across all OCR passes — most common valid result wins."""
    votes = [_parse_name_single(t) for t in all_texts]
    votes = [v for v in votes if v and all(len(w) >= 2 for w in v.split())]
    if not votes: return ""
    return Counter(votes).most_common(1)[0][0]

def get_aadhaar_phone(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    # Standalone 10-digit line (how Aadhaar prints it — alone at end of address block)
    for line in lines:
        clean = re.sub(r"[\s\-]", "", line).replace("$", "9")
        if re.fullmatch(r"[6-9]\d{9}", clean):
            return clean
    # Labeled or anywhere
    m = re.search(r"(?:mobile|phone|mob|cell|ph|tel)[:\s#\-]*([6-9]\d{9})", text, re.IGNORECASE)
    if m: return m.group(1)
    m = re.search(r"\b([6-9]\d{9})\b", text)
    return m.group(1) if m else ""

def get_aadhaar_address(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    # Find start anchor: D/O line or house # line
    start = None
    for i, line in enumerate(lines):
        if re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", line, re.IGNORECASE):
            start = i; break
    if start is None:
        for i, line in enumerate(lines):
            if re.match(r"^#\s*\d+", line):
                start = i; break
    # Fallback: grab lines around PIN code
    if start is None:
        for i, line in enumerate(lines):
            if re.search(r"\b[1-9]\d{5}\b", line):
                return ", ".join(lines[max(0, i - 6): i + 1])
        return ""
    # Collect from anchor until we hit the 6-digit PIN line
    addr = []
    for line in lines[start:]:
        addr.append(line)
        if re.search(r"\b[1-9]\d{5}\b", line):
            break
    # Remove D/O, S/O, W/O, C/O line from final address
    addr = [line for line in addr if not re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", line, re.IGNORECASE)]
    return ", ".join(addr)


# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────

def to_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws    = writer.sheets["Data"]
        fill  = PatternFill("solid", fgColor="1565C0")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return buf.getvalue()


# ──────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────

col1, col2 = st.columns(2)

with col1:
    st.subheader("📋 Upload Resumes")
    st.caption("Format: PDF")
    resume_files = st.file_uploader("Resume PDFs", type=["pdf"],
                                    accept_multiple_files=True, key="res",
                                    label_visibility="collapsed")

with col2:
    st.subheader("🪪 Upload Aadhaar Cards")
    st.caption("Formats: PDF · JPG · PNG")
    aadhaar_files = st.file_uploader("Aadhaar files", type=["pdf", "jpg", "jpeg", "png"],
                                     accept_multiple_files=True, key="aad",
                                     label_visibility="collapsed")

if st.button("🚀 Extract", use_container_width=True):

    if not resume_files and not aadhaar_files:
        st.warning("Please upload at least one file.")
        st.stop()

    # ── Step 1: Extract Aadhaar records ──
    aadhaar_records = []
    if aadhaar_files:
        bar_aad = st.progress(0, text="Extracting Aadhaar files…")
        for i, f in enumerate(aadhaar_files):
            raw = f.read()
            ext = f.name.lower().rsplit(".", 1)[-1]
            if ext == "pdf":
                primary   = pdf_to_text(raw)
                all_texts = [primary]
            else:
                primary, all_texts = aadhaar_multipass_ocr(raw)
            aadhaar_records.append({
                "aadhaar_name":   get_aadhaar_name(all_texts),
                "aadhaar_mobile": get_aadhaar_phone(primary),
                "address":        get_aadhaar_address(primary),
            })
            bar_aad.progress((i + 1) / len(aadhaar_files))
        bar_aad.empty()

    # ── Step 2: Extract Resume records, match to Aadhaar by name similarity ──
    from difflib import SequenceMatcher

    def name_sim(a, b):
        if not a or not b: return 0.0
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()

    def match(resume_name):
        if not aadhaar_records: return {}
        best, best_score = {}, 0.0
        for rec in aadhaar_records:
            s = name_sim(resume_name, rec["aadhaar_name"])
            if s > best_score: best_score, best = s, rec
        return best if best_score > 0.35 else {}

    rows = []
    if resume_files:
        bar_res = st.progress(0, text="Extracting Resume files…")
        for i, f in enumerate(resume_files):
            text   = pdf_to_text(f.read())
            rname  = get_resume_name(text)
            matched = match(rname)
            rows.append({
                "Resume Name":    rname,
                "Email":          get_resume_email(text),
                "Resume Mobile":  get_resume_phone(text),
                "College":        get_resume_college(text),
                "Aadhaar Name":   matched.get("aadhaar_name", ""),
                "Aadhaar Mobile": matched.get("aadhaar_mobile", ""),
                "Address":        matched.get("address", ""),
            })
            bar_res.progress((i + 1) / len(resume_files))
        bar_res.empty()
    elif aadhaar_records:
        # Only Aadhaar uploaded — still show combined columns with blanks for resume fields
        for rec in aadhaar_records:
            rows.append({
                "Resume Name":    "",
                "Email":          "",
                "Resume Mobile":  "",
                "College":        "",
                "Aadhaar Name":   rec["aadhaar_name"],
                "Aadhaar Mobile": rec["aadhaar_mobile"],
                "Address":        rec["address"],
            })

    # ── Step 3: Show combined table ──
    COLS = ["Resume Name", "Email", "Resume Mobile", "College",
            "Aadhaar Name", "Aadhaar Mobile", "Address"]
    df = pd.DataFrame(rows, columns=COLS)

    st.subheader("📊 Extracted Data")
    st.dataframe(df, use_container_width=True, height=400)
    st.download_button(
        "⬇️ Download Excel", to_excel(df),
        file_name="resume_aadhaar_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
