"""
DocMatch — Resume & Aadhaar Extractor
Desktop app (no Streamlit, no extra installs beyond standard Python libs + pdfplumber/pytesseract/openpyxl).

Run:  python DocMatch.pyw   (or double-click on Windows with pythonw)

Required pip packages (one-time):
    pip install pdfplumber pdf2image pillow pytesseract openpyxl pandas

Tesseract-OCR must be installed on the OS (https://github.com/UB-Mannheim/tesseract/wiki for Windows).
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import io
import re
import os
import sys
import platform
import json
import urllib.request
from collections import Counter
from difflib import SequenceMatcher

# ── Graceful imports ──────────────────────────────────────────────────────────
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from pdf2image import convert_from_bytes
except ImportError:
    convert_from_bytes = None

try:
    from PIL import Image, ImageEnhance, ImageTk
except ImportError:
    Image = ImageEnhance = ImageTk = None

try:
    import pytesseract
except ImportError:
    pytesseract = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    openpyxl = None

# ── Colours & fonts ───────────────────────────────────────────────────────────
BG          = "#F4F6FA"
DARK        = "#1A2340"
CARD        = "#FFFFFF"
ACCENT      = "#3B4FD4"
ACCENT2     = "#5B6FF4"
BORDER      = "#D5DAE8"
TEXT        = "#1E2A45"
SUBTEXT     = "#6B748A"
SUCCESS     = "#16A34A"
TAG_BG      = "#EEF1FB"
TAG_FG      = "#3B4FD4"
ROW_ALT     = "#F8F9FD"
HEADER_BG   = "#F0F2FA"

FONT_HEAD   = ("Segoe UI", 13, "bold")
FONT_SUB    = ("Segoe UI", 9)
FONT_BODY   = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 8)
FONT_BTN    = ("Segoe UI", 11, "bold")
FONT_TITLE  = ("Segoe UI", 14, "bold")
FONT_NAV    = ("Segoe UI", 11, "bold")


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION LOGIC (identical to your Streamlit version)
# ─────────────────────────────────────────────────────────────────────────────

def pdf_to_text(file_bytes):
    text = ""
    if pdfplumber:
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
        except Exception:
            pass
    if len(text.strip()) < 30 and convert_from_bytes and pytesseract and Image:
        try:
            for img in convert_from_bytes(file_bytes, dpi=300):
                gray = ImageEnhance.Contrast(img.convert("L")).enhance(2.0)
                text += pytesseract.image_to_string(gray) + "\n"
        except Exception:
            pass
    return text


def aadhaar_multipass_ocr(file_bytes):
    if not (Image and ImageEnhance and pytesseract):
        return "", []
    img = Image.open(io.BytesIO(file_bytes))
    w, h = img.size
    results = []
    for crop_r, contrast in [(0.60, 2.5), (0.60, 3.0), (0.62, 2.5), (0.58, 2.5), (0.60, 2.0)]:
        crop = img.crop((0, 0, w, int(h * crop_r)))
        up   = crop.resize((w * 3, int(h * crop_r) * 3), Image.LANCZOS)
        gray = ImageEnhance.Contrast(up.convert("L")).enhance(contrast)
        results.append(pytesseract.image_to_string(gray, config="--oem 3 --psm 6"))
    return results[0], results


def get_resume_name(text):
    skip = {"resume", "cv", "curriculum vitae", "profile", "summary",
            "objective", "contact", "email", "phone", "mobile", "address", "name"}
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines[:12]:
        if any(k in line.lower() for k in skip): continue
        if re.search(r"[^a-zA-Z\s\.\-]", line): continue
        if 2 <= len(line.split()) <= 5 and re.match(r"^[A-Z][a-z]+(\s[A-Z][a-z]+)+$", line):
            return line
    for line in lines[:8]:
        if 2 <= len(line.split()) <= 5 and re.match(r"^[A-Za-z\s\.\-]+$", line):
            return line.title()
    return ""


def get_resume_email(text):
    m = re.search(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return m.group(0) if m else ""


def get_resume_phone(text):
    def clean(raw):
        digits = re.sub(r"[^\d]", "", raw)
        if len(digits) == 12 and digits.startswith("91"):   digits = digits[2:]
        elif len(digits) == 13 and digits.startswith("091"): digits = digits[3:]
        elif len(digits) == 14 and digits.startswith("0091"): digits = digits[4:]
        elif len(digits) == 11 and digits.startswith("0"):   digits = digits[1:]
        if len(digits) == 10 and digits[0] in "6789": return digits
        return ""

    patterns = [
        r"(?:phone|mobile|mob|cell|ph|tel|contact|whatsapp)[^\d]{0,5}(\+?0{0,2}91[\s\-\.]?[6-9][\d\s\-\.]{9,14})",
        r"(?:phone|mobile|mob|cell|ph|tel|contact|whatsapp)[^\d]{0,5}(0?[6-9][\d\s\-\.]{9,13})",
        r"(\+91[\s\-\.]?[6-9][\d\s\-\.]{9,12})",
        r"(0091[\s\-\.]?[6-9][\d\s\-\.]{9,12})",
        r"\b(0[6-9]\d{9})\b",
        r"\b([6-9]\d{4}[\s\-\.]\d{5})\b",
        r"\b([6-9]\d{3}[\s\-\.]\d{6})\b",
        r"\b([6-9]\d{2}[\s\-\.]\d{7})\b",
        r"\(([6-9]\d{4})\)[\s\-\.]?(\d{5})",
        r"\b([6-9]\d{9})\b",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            raw    = "".join(g for g in m.groups() if g)
            result = clean(raw)
            if result: return result
    return ""



def get_resume_qualification(text):
    """Extracts degree/qualification: B.Tech, BCA, MBA, M.Tech, B.Com etc."""
    DEGREES = [
        (r"\bph\.?\s*d\.?\b",               "Ph.D"),
        (r"\bm\.?\s*tech\b",                  "M.Tech"),
        (r"\bm\.?\s*e\.?\b",                 "M.E"),
        (r"\bm\.?\s*sc\.?\b",                "M.Sc"),
        (r"\bm\.?\s*com\.?\b",               "M.Com"),
        (r"\bm\.?\s*ca\b",                    "MCA"),
        (r"\bm\.?\s*ba\b",                    "MBA"),
        (r"\bm\.?\s*arch\.?\b",              "M.Arch"),
        (r"\bm\.?\s*pharm\.?\b",             "M.Pharm"),
        (r"\bmaster\s+of\s+[a-z\s]{3,30}",    None),
        (r"\bb\.?\s*tech\b",                  "B.Tech"),
        (r"\bb\.?\s*e\.?\b",                 "B.E"),
        (r"\bb\.?\s*sc\.?\b",                "B.Sc"),
        (r"\bb\.?\s*com\.?\b",               "B.Com"),
        (r"\bb\.?\s*ca\b",                    "BCA"),
        (r"\bb\.?\s*ba\b",                    "BBA"),
        (r"\bb\.?\s*arch\.?\b",              "B.Arch"),
        (r"\bb\.?\s*pharm\.?\b",             "B.Pharm"),
        (r"\bbachelor\s+of\s+[a-z\s]{3,30}",  None),
        (r"\bdiploma\b",                        "Diploma"),
    ]
    for pattern, canonical in DEGREES:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            if canonical:
                return canonical
            raw = m.group(0).strip()
            raw = re.split(r"[,|\n(]", raw)[0].strip()
            return raw.title() if len(raw) < 35 else ""
    return ""

_COLLEGE_WORDS_RE = r"(college|university|institute|of|engineering|technology|science|arts|commerce|school|academy|polytechnic|management)"

def _split_camel_college(name):
    """
    Fix CamelCase / run-together college names.
    'KORMcollegeofEngineering'  → 'KORM college of Engineering'
    'RVCollegeofEngineering'    → 'RV College of Engineering'
    'ChristUniversity,Bangalore'→ 'Christ University, Bangalore'
    """
    # Split before known education words glued to preceding text
    name = re.sub(rf"(?<=[a-zA-Z])({_COLLEGE_WORDS_RE})", r" \1", name, flags=re.IGNORECASE)
    # Standard camelCase: lowercase → uppercase
    name = re.sub(r"([a-z])([A-Z])", r"\1 \2", name)
    # ACRONYM → lowercase (e.g. "KORMc" → "KORM c"), but not trailing 's'
    name = re.sub(r"([A-Z]{2,})([a-z])",
                  lambda m: m.group(1) + " " + m.group(2) if m.group(2) != "s" else m.group(0),
                  name)
    # Fix missing space after comma
    name = re.sub(r",([^\s])", r", \1", name)
    # Collapse spaces
    name = re.sub(r"\s{2,}", " ", name)
    return name.strip()


def _title_college(name):
    """
    Smart title-case that keeps known acronyms upper and small words (of/and/the) lower.
    'OSMANIA UNIVERSITY' → 'Osmania University'
    'iit bombay'         → 'IIT Bombay'
    """
    ACRONYMS = {
        "iit", "nit", "iiit", "bits", "mit", "vit", "srm", "lpu", "kl",
        "jntu", "jntuk", "jntuh", "jntua", "ktu", "apj", "rgpv", "csvtu",
        "mgu", "kuk", "du", "bhu", "hcu", "uoh", "au", "ou", "svu",
        "rv", "korm", "ugc", "aicte", "naac",
    }
    SMALL = {"of", "and", "the", "in", "at", "for", "to", "a", "an"}
    words = name.split()
    result = []
    for i, w in enumerate(words):
        wl = w.lower().strip(".,")
        if wl in ACRONYMS:
            result.append(wl.upper())
        elif wl in SMALL and i > 0:
            result.append(wl)
        else:
            result.append(w.capitalize())
    return " ".join(result)


def _strip_degree_and_branch(line):
    """Strip degree keyword + optional 'in BRANCH (spec)' from start of line."""
    # Pass A: strip degree keyword at the start
    line = re.sub(
        r"^\s*(b\.?\s*tech|b\.?\s*e\.?|b\.?\s*sc\.?|m\.?\s*tech|m\.?\s*sc\.?|mba|mca|bca|bba|"
        r"phd|ph\.d|bachelor[s]?\s*(?:of\s*\w+)?|master[s]?\s*(?:of\s*\w+)?|diploma|be\\b|me\\b|"
        r"m\.e|b\.arch|m\.arch)\s*",
        "", line, flags=re.IGNORECASE
    )
    # Pass B: strip remaining "in BRANCH (optional specialisation)" prefix
    line = re.sub(
        r"^\s*in\s+[\w\s/&\.\-]+(?:\([^)]*\))?\s*[-–—,:/]?\s*",
        "", line, flags=re.IGNORECASE
    )
    return line.strip()


def _clean_college_line(line):
    """10-step pipeline to produce a clean college name from any raw resume line."""
    line = line.strip()
    # 1. Bullets
    line = re.sub(r"^[\+\-\–\—\•\*\·\>\|▪►▶]+\s*", "", line)
    # 2. Degree + branch prefix (two clean passes)
    line = _strip_degree_and_branch(line)
    # 3. Remaining branch/stream prefix e.g. "CSE, RV College"
    BRANCHES = (
        r"computer science(\s*and\s*engineering)?|information technology|"
        r"electronics(\s*and\s*communication)?|electrical(\s*and\s*electronics)?|"
        r"mechanical engineering|civil engineering|chemical engineering|"
        r"artificial intelligence(\s*(and\s*(machine learning|data science))?)?|"
        r"data science|cyber security|biotechnology|cse|ece|eee|mech|civil"
    )
    line = re.sub(rf"^\s*({BRANCHES})\s*[,\-–]?\s*", "", line, flags=re.IGNORECASE)
    # 4. Pipe: keep college part
    if "|" in line:
        STRONG = ["university", "college", "institute", "iit", "nit", "iiit",
                  "bits", "polytechnic", "academy"]
        for part in line.split("|"):
            if any(k in part.lower() for k in STRONG):
                line = part.strip()
                break
    # 5. Affiliated to
    line = re.sub(r"\s*(affiliated\s*(to|with|under)|under\s+the\s+affiliation\s+of).*$",
                  "", line, flags=re.IGNORECASE)
    # 6. Year ranges — handles "2020-2023", "Jan 2021 - May 2024", "Aug 2018 - Present"
    _MONTHS = r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s*"
    line = re.sub(
        rf"[\s,|]+(?:{_MONTHS})?\d{{4}}\s*[-–—]\s*(?:{_MONTHS})?(?:\d{{4}}|present|current|ongoing|till\s*date).*$",
        "", line, flags=re.IGNORECASE
    )
    line = re.sub(r"[\s|]+\d{{4}}\s*$", "", line)
    # 7. CGPA / grade noise
    line = re.sub(
        r"[\|,;\s]+(cgpa|gpa|percentage|%|grade|marks|score|pass|distinction|"
        r"first class|second class|honours|honors|aggregate).*$",
        "", line, flags=re.IGNORECASE
    )
    # 8. CamelCase fix
    line = _split_camel_college(line)
    # 9. Whitespace + smart title-case
    line = re.sub(r"\s{2,}", " ", line).strip(" ,;|–-—")
    if line == line.upper() and len(line) > 4:
        line = _title_college(line)
    elif line == line.lower() and len(line) > 4:
        line = _title_college(line)
    # 10. Trim trailing city/place noise not separated by comma
    _INST_END = {"engineering", "technology", "university", "college", "institute",
                 "polytechnic", "academy", "sciences", "management", "arts", "commerce", "school"}
    words = line.split()
    last_idx = -1
    for i, w in enumerate(words):
        if w.lower().rstrip(".,") in _INST_END:
            last_idx = i
    if last_idx >= 0 and last_idx < len(words) - 1:
        before_trailing = line[:line.rfind(" ".join(words[last_idx + 1:]))].rstrip()
        if "," not in before_trailing:
            line = " ".join(words[:last_idx + 1])
    return line



# ── Gemini API key (free, no billing needed) ─────────────────────────────────
# Get a free key at: https://aistudio.google.com/apikey
GEMINI_API_KEY = ""   # ← paste your free key here

def _ai_extract_college(resume_text):
    """
    Uses Google Gemini (free API) to extract the college name from resume text.
    Falls back silently if key is missing or call fails.
    """
    if not GEMINI_API_KEY:
        return ""
    try:
        # Take only first 3000 chars to keep prompt small and fast
        snippet = resume_text[:3000]
        prompt = (
            "Extract ONLY the most recent college/university/institute name from this resume text. "
            "Return just the institution name, nothing else — no degree, no year, no branch, no location. "
            "If not found, return empty string.\n\nResume:\n" + snippet
        )
        payload = json.dumps({
            "contents": [{"parts": [{"text": prompt}]}]
        }).encode("utf-8")
        url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
               f"gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}")
        req = urllib.request.Request(
            url, data=payload,
            headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        result = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        # Sanity check: should look like a college name
        if len(result) < 4 or len(result) > 120:
            return ""
        if re.search(r"[\n\r]", result):   # reject multi-line responses
            result = result.splitlines()[0].strip()
        return result
    except Exception:
        return ""


def get_resume_college(text):
    STRONG_KW = ["university", "college", "institute", "institution", "deemed"]
    BROAD_KW  = ["iit", "nit", "iiit", "iim", "xlri", "bits",
                 "polytechnic", "engineering", "academy", "school of"]

    # Known short-form institutions (no standard keyword in name)
    SHORT_INST = re.compile(
        r"^\s*(IIT|NIT|IIIT|IIM|XLRI|BITS|VIT|SRM|LPU|TISS|NIFT|NID|MICA)\b",
        re.IGNORECASE)

    # Table-header words — skip these lines
    HEADER_WORDS = {"degree","year","score","grade","marks","cgpa",
                    "percentage","board","qualification","institution","stream"}

    def split_camel(s):
        edu = r"(college|university|institute|engineering|technology|school|academy|polytechnic|management|science|commerce|arts)"
        s = re.sub(rf"(?<=[a-zA-Z])({edu})", r" \1", s, flags=re.IGNORECASE)
        s = re.sub(r"([A-Z]{2,})([a-z])",       r"\1 \2", s)
        s = re.sub(r"([a-z])([A-Z])",            r"\1 \2", s)
        s = re.sub(r"([A-Z]{2,})([A-Z][a-z])",   r"\1 \2", s)
        s = re.sub(r"(college|university|institute|school|academy)(of|and|for|the)",
                   r"\1 \2", s, flags=re.IGNORECASE)
        s = re.sub(r"\b(of|and|for|the)([A-Z])", r"\1 \2", s)
        s = re.sub(r",([^\s])", r", \1", s)
        s = re.sub(r"\s{2,}", " ", s)
        return s

    def clean_college(line):
        # 0. Fix OCR pipe noise inside words e.g. "col|ege" → "college"
        line = re.sub(r"(?<=[a-zA-Z])\|(?=[a-zA-Z])", "", line)
        line = split_camel(line)

        # 1. Strip leading bullets: • + * - | etc.
        line = re.sub(r"^[\s\+\•\*\-\–\—\►\▪\·\|]+", "", line)

        # 2. Strip brackets early so "(AI and ML)" is gone before branch check
        line = re.sub(r"\s*[\(\[].*?[\)\]]", "", line)

        # 3. Strip branch prefix e.g. "CSE," or "in CSE"
        BRANCHES = (r"CSE|ECE|EEE|EIE|ICE|IT\b|CS\b|MECH|CIVIL|AERO|AUTO|"
                    r"AIDS|AIML|AI\b|ML\b|DS\b|VLSI|CHEM|BIOTECH|"
                    r"BCA|MCA|MBA|BBA|MIM")
        line = re.sub(rf"^(in\s*)?({BRANCHES})[,\s]+", "", line, flags=re.IGNORECASE)
        line = re.sub(r"^(in|at|from|of|the)\s+", "", line, flags=re.IGNORECASE)

        # 4. Strip degree prefix
        line = re.sub(
            r"^(bachelor\s+of\s+(computer\s+applications|engineering|technology|"
            r"science|commerce|arts)|master\s+of\s+(technology|engineering|science|"
            r"business\s+administration|computer\s+applications)|"
            r"b\.?\s*tech|b\.?\s*e\.?|m\.?\s*tech|m\.?\s*e\.?|"
            r"b\.?\s*sc\.?|m\.?\s*sc\.?|b\.?\s*com|m\.?\s*com|"
            r"b\.?\s*ca|m\.?\s*ca|b\.?\s*ba|m\.?\s*ba|"
            r"ph\.?\s*d\.?|diploma|bachelor|master|"
            r"intermediate|pre[\s\-]?university|p\.?\s*u\.?\s*c\.?|"
            r"10\+2|hsc|ssc|secondary|higher\s+secondary)"
            r"\s*[\-–,:/\(]?\s*",
            "", line, flags=re.IGNORECASE)

        # 5. Strip "in BRANCH" anywhere in line (may appear after degree removal)
        line = re.sub(
            rf"\bin\s+({BRANCHES})\b[,\s]*", " ", line, flags=re.IGNORECASE)

        # 5b. Strip full branch names e.g. "Computer Science and Engineering -"
        line = re.sub(
            r"^(computer\s+science(\s+and\s+engineering)?|"
            r"information\s+technology|"
            r"electronics(\s+and\s+communication(\s+engineering)?)?|"
            r"electrical(\s+and\s+electronics(\s+engineering)?)?|"
            r"mechanical\s+engineering|civil\s+engineering|"
            r"chemical\s+engineering|"
            r"artificial\s+intelligence(\s+and\s+machine\s+learning)?|"
            r"data\s+science|cyber\s+security|biotechnology)"
            r"\s*[\-–,]?\s*",
            "", line, flags=re.IGNORECASE)

        line = re.sub(r"\s{2,}", " ", line).strip()

        # 6. If too many junk words before first institution keyword, trim them
        INST_KW = {"university","college","institute","institution",
                   "polytechnic","academy","school","engineering","technology"}
        words = line.split()
        first_inst = next((i for i, w in enumerate(words)
                           if w.lower().rstrip(",.") in INST_KW), -1)
        if first_inst > 4:
            line = " ".join(words[max(0, first_inst - 2):])

        # 7. Pipe-separated: keep the part with institution keyword
        if "|" in line:
            for part in line.split("|"):
                p = part.strip()
                if any(k in p.lower() for k in STRONG_KW + ["iit","nit","bits","iim"]):
                    line = p
                    break

        # 8. Strip affiliated / board names
        line = re.sub(r"[,\s]*affiliated\s*(to|by|under)?.*$",
                      "", line, flags=re.IGNORECASE)
        line = re.sub(
            r",?\s*\b(APOSS|CBSE|ICSE|IGCSE|SSC|HSC|BIEC|BISE|NIOS|CISCE|"
            r"state\s*board|central\s*board|telangana\s*board|"
            r"andhra\s*board|board\s*of\s*\w+)\b.*$",
            "", line, flags=re.IGNORECASE)

        # 9. Strip date ranges: "Jan-May", "Jan 2021 - May 2024", "2020-2023"
        line = re.sub(
            r"\s*\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*"
            r"(\s+\d{4})?\s*[-–]\s*"
            r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*(\s+\d{4})?\b",
            "", line, flags=re.IGNORECASE)
        line = re.sub(
            r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}\b",
            "", line, flags=re.IGNORECASE)
        line = re.sub(r"\b\d{4}\s*[-–]\s*(present|\d{4})\b",
                      "", line, flags=re.IGNORECASE)
        line = re.sub(r"\b\d{4}\b", "", line)

        # 10. Strip CGPA / percentage / grade
        line = re.sub(r"\s*(cgpa|gpa|percentage|marks|grade)\b.*$",
                      "", line, flags=re.IGNORECASE)
        line = re.sub(r"[\s,\-]*\d+\.?\d*\s*%.*$", "", line)

        # 11. Strip trailing city after comma (1-2 word location)
        m = re.match(r"^(.+?),\s*([A-Z][a-zA-Z\s]{2,20})$", line)
        if m:
            college_part = m.group(1).strip()
            after_words  = m.group(2).strip().lower().split()
            KEEP = {"of","and","for","the","engineering","technology",
                    "science","arts","commerce","management","higher","education"}
            first = college_part.split()[0].lower() if college_part else ""
            KEEP_COMMA_INST = {"iit","nit","iiit","bits","iim","xlri","mit","vit","srm","lpu"}
            if first not in KEEP_COMMA_INST and \
               not any(w in KEEP for w in after_words) and len(after_words) <= 2:
                line = college_part

        # 12. Final cleanup
        line = re.sub(r"[^a-zA-Z\s,\.&]", "", line)
        line = re.sub(r"\s{2,}", " ", line).strip().strip(",-. ").strip()
        if line and line == line.upper() and len(line) > 4:
            line = line.title()
        return line

    def is_header(line):
        """Skip table header rows like 'Degree | Institution | Year | %'"""
        words = set(re.sub(r"[\|\s]+", " ", line).strip().lower().split())
        return len(words) <= 7 and len(words & HEADER_WORDS) >= 2

    # ── Pass 1: lines with strong institution keywords ────────────────────────
    candidates = []
    for line in text.splitlines():
        low = line.lower()
        if any(k in low for k in STRONG_KW):
            if is_header(line):
                continue
            c = clean_college(line)
            if 4 < len(c) < 120:
                candidates.append(c)

    if candidates:
        # Return the FIRST valid candidate (most recent = appears first in resume)
        return candidates[0]

    # ── Pass 2: college name on NEXT LINE after degree keyword ───────────────
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    DEGREE_PAT = re.compile(
        r"\b(b\.?tech|b\.?e\b|m\.?tech|bca|mba|mca|bsc|msc|diploma|"
        r"bachelor|master|intermediate)\b", re.IGNORECASE)
    for i, line in enumerate(lines):
        if DEGREE_PAT.search(line):
            for nxt in lines[i+1:i+4]:
                c = clean_college(nxt)
                if (4 < len(c) < 120 and
                        any(k in c.lower() for k in STRONG_KW + ["iit","nit","bits","iim"])):
                    return c

    # ── Pass 3: short known institutions e.g. "IIT Delhi" ────────────────────
    for line in text.splitlines():
        if SHORT_INST.search(line):
            c = clean_college(line)
            if 4 < len(c) < 80:
                return c

    # ── Pass 4: broad keywords ────────────────────────────────────────────────
    for line in text.splitlines():
        if any(k in line.lower() for k in BROAD_KW):
            c = clean_college(line)
            if 4 < len(c) < 120:
                return c

    return _ai_extract_college(text)


def _is_name(s):
    s = s.strip()
    if len(s) < 4: return False
    words = s.split()
    if not (2 <= len(words) <= 6): return False
    return all(re.match(r"^[A-Za-z\.\-]+$", w) for w in words)


def _parse_name_single(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for i, line in enumerate(lines):
        if re.fullmatch(r"To\.?", line, re.IGNORECASE):
            cands, cand_idx = [], []
            for j in range(i + 1, min(i + 9, len(lines))):
                c = re.sub(r"^[\-\—\s]+|[\-\—\s]+$", "", lines[j]).strip()
                if re.match(r"^(D/O|S/O|W/O|C/O|O:|#\d)", c, re.IGNORECASE): break
                if _is_name(c): cands.append(c.title()); cand_idx.append(j)
            if len(cands) >= 2: return cands[1]
            if len(cands) == 1:
                nxt = cand_idx[0] + 1
                if nxt < len(lines) and re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", lines[nxt], re.IGNORECASE):
                    return cands[0]
    for i, line in enumerate(lines):
        if re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", line, re.IGNORECASE):
            for j in range(i - 1, max(i - 4, -1), -1):
                c = re.sub(r"^[\-\—\s]+|[\-\—\s]+$", "", lines[j]).strip()
                if _is_name(c): return c.title()
    for i, line in enumerate(lines):
        if re.search(r"\bDOB\b|\bDate of Birth\b|\bD\.O\.B\b", line, re.IGNORECASE):
            for j in range(i - 1, max(i - 4, -1), -1):
                c = re.sub(r"^[\-\—\s]+|[\-\—\s]+$", "", lines[j]).strip()
                if _is_name(c): return c.title()
    return ""


def get_aadhaar_name(all_texts):
    votes = [_parse_name_single(t) for t in all_texts]
    votes = [v for v in votes if v and all(len(w) >= 2 for w in v.split())]
    if not votes: return ""
    return Counter(votes).most_common(1)[0][0]


def get_aadhaar_phone(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for line in lines:
        clean = re.sub(r"[\s\-]", "", line).replace("$", "9")
        if re.fullmatch(r"[6-9]\d{9}", clean): return clean
    m = re.search(r"(?:mobile|phone|mob|cell|ph|tel)[:\s#\-]*([6-9]\d{9})", text, re.IGNORECASE)
    if m: return m.group(1)
    m = re.search(r"\b([6-9]\d{9})\b", text)
    return m.group(1) if m else ""


def _clean_address_parts(parts):
    """
    Clean a list of address line strings:
    - Remove OCR noise lines (too short, all symbols, random chars)
    - Remove lines that are clearly non-address (dates, Aadhaar number, names, DOB)
    - Normalise spacing and punctuation
    """
    cleaned = []
    for p in parts:
        p = p.strip().strip(",;|")
        p = re.sub(r"\s{2,}", " ", p)

        if not p or len(p) < 3:
            continue
        # Skip lines that look like Aadhaar number (12 digits)
        if re.fullmatch(r"[\d\s\-]{12,16}", p):
            continue
        # Skip DOB lines
        if re.search(r"\b(DOB|Date of Birth|D\.O\.B|Year of Birth)\b", p, re.IGNORECASE):
            continue
        # Skip lines that are only symbols / garbage (no letters or digits)
        if not re.search(r"[A-Za-z0-9]", p):
            continue
        # Skip very short lines that are just 1-2 chars (OCR noise)
        if len(p) <= 2:
            continue
        # Skip lines starting with known non-address prefixes
        if re.match(r"^(Government of India|Unique Identification|UIDAI|आधार|Aadhaar|VID\s*:)", p, re.IGNORECASE):
            continue
        # Clean up common OCR glitches at start of parts
        p = re.sub(r"^[,;|\-–—]+\s*", "", p)
        p = re.sub(r"\s*[,;|\-–—]+$", "", p)
        if p:
            cleaned.append(p)
    return cleaned


def get_aadhaar_address(text):
    """
    Extracts and cleans the address from Aadhaar OCR text.
    Supports formats:
      - Standard: D/O ... #House No, Street, Village, District, State - PIN
      - #2019, 6th cross samadhi colony, Bengaluru, Karnataka - 560023
      - Flat/Door No + Street + Landmark + City + State + PIN (multi-line)
      - Single-line compact address
    Returns a clean, comma-separated address string ending at PIN code.
    """
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # ── Find start anchor ──
    start = None
    for i, line in enumerate(lines):
        if re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", line, re.IGNORECASE):
            start = i; break
    if start is None:
        for i, line in enumerate(lines):
            if re.match(r"^#\s*\d+", line):
                start = i; break
    if start is None:
        # Try to find a house/flat/door number pattern
        for i, line in enumerate(lines):
            if re.match(r"^(flat|door|house|plot|h\.?no|d\.?no|f\.?no|s\.?no|no\.?)\s*[:\-\.]?\s*\d",
                        line, re.IGNORECASE):
                start = i; break
    if start is None:
        # Fallback: grab lines around 6-digit PIN
        for i, line in enumerate(lines):
            if re.search(r"\b[1-9]\d{5}\b", line):
                raw_parts = []
                for ln in lines[max(0, i - 6): i + 1]:
                    raw_parts.append(ln)
                cleaned = _clean_address_parts(raw_parts)
                return ", ".join(cleaned)
        return ""

    # ── Collect lines from anchor to PIN ──
    raw_addr = []
    for line in lines[start:]:
        raw_addr.append(line)
        if re.search(r"\b[1-9]\d{5}\b", line):
            break
        # Safety: stop after 10 lines to avoid grabbing entire doc
        if len(raw_addr) > 10:
            break

    # ── Remove D/O, S/O, W/O, C/O relation lines ──
    raw_addr = [l for l in raw_addr
                if not re.match(r"^(D/O|S/O|W/O|C/O|O:)[:\s]", l, re.IGNORECASE)]

    # ── Clean individual parts ──
    cleaned = _clean_address_parts(raw_addr)

    # ── Final formatting: join with ", " and ensure PIN is at end ──
    # Sometimes the state and PIN are on separate lines — merge them
    result_parts = []
    i = 0
    while i < len(cleaned):
        part = cleaned[i]
        # If this part ends with a state name and next part is just a PIN, merge
        if i + 1 < len(cleaned) and re.fullmatch(r"[1-9]\d{5}", cleaned[i + 1].strip()):
            result_parts.append(part + " - " + cleaned[i + 1].strip())
            i += 2
            continue
        # If PIN is inline (e.g. "Karnataka - 560023" or "Karnataka 560023")
        part = re.sub(r"\s*[-–]\s*([1-9]\d{5})\s*$", r" - \1", part)
        result_parts.append(part)
        i += 1

    return ", ".join(result_parts)


def to_excel_bytes(rows):
    if not openpyxl or not pd:
        return None
    COLS = ["Resume Name", "Email", "Resume Mobile", "Qualification", "College",
            "Aadhaar Name", "Aadhaar Mobile", "Address"]
    df  = pd.DataFrame(rows, columns=COLS)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws   = writer.sheets["Data"]
        fill = PatternFill("solid", fgColor="1565C0")
        hfnt = Font(bold=True, color="FFFFFF", size=11)
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)
        for cell in ws[1]:
            cell.fill = fill; cell.font = hfnt
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return buf.getvalue()


def name_sim(a, b):
    if not a or not b: return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────

class DocMatchApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("DocMatch — Resume & Aadhaar Extractor")
        self.geometry("1100x720")
        self.minsize(900, 600)
        self.configure(bg=BG)
        self.resizable(True, True)

        # State
        self.resume_files  = []   # list of (name, bytes)
        self.aadhaar_files = []   # list of (name, bytes)
        self.rows          = []   # extracted rows for table / excel

        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x  = (sw - 1100) // 2
        y  = (sh - 720)  // 2
        self.geometry(f"1100x720+{x}+{y}")

    # ── Build UI ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── NAV BAR ──
        nav = tk.Frame(self, bg=DARK, height=52)
        nav.pack(fill="x", side="top")
        nav.pack_propagate(False)

        logo_box = tk.Frame(nav, bg="#3B4FD4", width=36, height=36)
        logo_box.place(relx=0, rely=0.5, anchor="w", x=16, y=0)
        tk.Label(logo_box, text="D", font=("Segoe UI", 14, "bold"),
                 fg="white", bg="#3B4FD4").place(relx=0.5, rely=0.5, anchor="center")

        tk.Label(nav, text="DocMatch", font=FONT_NAV, fg="white", bg=DARK
                 ).place(relx=0, rely=0.5, anchor="w", x=60)
        tk.Label(nav, text="Resume & Aadhaar Extractor", font=FONT_SUB,
                 fg="#8A93B0", bg=DARK
                 ).place(relx=0, rely=0.5, anchor="w", x=145)

        # ── MAIN SCROLL AREA ──
        self.main = tk.Frame(self, bg=BG)
        self.main.pack(fill="both", expand=True, padx=24, pady=18)

        # ── UPLOAD SECTION LABEL ──
        tk.Label(self.main, text="UPLOAD DOCUMENTS", font=("Segoe UI", 9, "bold"),
                 fg=SUBTEXT, bg=BG).pack(anchor="w", pady=(0, 8))

        # ── UPLOAD CARDS ROW ──
        cards_row = tk.Frame(self.main, bg=BG)
        cards_row.pack(fill="x")

        self.resume_card  = self._upload_card(cards_row, "📋", "Resume PDFs",
                                               "Drop files or click to browse",
                                               ["PDF"], self._browse_resumes, side="left")
        tk.Frame(cards_row, bg=BG, width=16).pack(side="left")
        self.aadhaar_card = self._upload_card(cards_row, "🪪", "Aadhaar Cards",
                                               "Drop files or click to browse",
                                               ["PDF", "JPG", "PNG"], self._browse_aadhaar, side="left")

        # File list labels (inside cards) — we store refs to update them
        # (already stored inside _upload_card as .file_frame)

        # ── EXTRACT BUTTON ──
        btn_frame = tk.Frame(self.main, bg=BG)
        btn_frame.pack(fill="x", pady=14)

        self.extract_btn = tk.Button(
            btn_frame, text="🚀  Extract Data",
            font=FONT_BTN, fg="white", bg=ACCENT,
            activebackground=ACCENT2, activeforeground="white",
            bd=0, cursor="hand2", relief="flat",
            command=self._start_extract,
            pady=14
        )
        self.extract_btn.pack(fill="x")
        self._bind_hover(self.extract_btn, ACCENT, ACCENT2)

        # ── PROGRESS ──
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.main, variable=self.progress_var,
                                             maximum=100, mode="determinate")
        self.status_lbl   = tk.Label(self.main, text="", font=FONT_SMALL,
                                      fg=SUBTEXT, bg=BG)

        # ── RESULTS SECTION ──
        res_header = tk.Frame(self.main, bg=BG)
        res_header.pack(fill="x", pady=(4, 0))

        self.result_title = tk.Label(res_header, text="📊  Extracted Data",
                                      font=FONT_TITLE, fg=TEXT, bg=BG)
        self.result_title.pack(side="left")

        self.record_lbl = tk.Label(res_header, text="", font=FONT_SUB,
                                    fg=SUBTEXT, bg=BG)
        self.record_lbl.pack(side="left", padx=(8, 0), pady=(4, 0))

        self.dl_btn = tk.Button(
            res_header, text="⬇  Download Excel",
            font=FONT_BODY, fg=ACCENT, bg=CARD,
            activebackground=TAG_BG, activeforeground=ACCENT,
            bd=1, relief="solid", cursor="hand2",
            command=self._download_excel,
            padx=12, pady=5
        )
        self.dl_btn.pack(side="right")
        self.dl_btn.config(state="disabled")

        # ── TABLE ──
        tbl_frame = tk.Frame(self.main, bg=CARD,
                              highlightbackground=BORDER, highlightthickness=1)
        tbl_frame.pack(fill="both", expand=True, pady=(8, 0))

        COLS = ("Resume Name", "Email", "Resume Mobile", "Qualification", "College",
                "Aadhaar Name", "Aadhaar Mobile", "Address")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                         background=CARD, fieldbackground=CARD,
                         foreground=TEXT, font=FONT_BODY,
                         rowheight=30)
        style.configure("Custom.Treeview.Heading",
                         background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", "#D6DCFB")],
                  foreground=[("selected", TEXT)])

        self.tree = ttk.Treeview(tbl_frame, columns=COLS, show="headings",
                                  style="Custom.Treeview")
        col_widths = [140, 180, 120, 110, 180, 140, 120, 220]
        for col, w in zip(COLS, col_widths):
            self.tree.heading(col, text=col.upper())
            self.tree.column(col, width=w, minwidth=80, anchor="w")

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tbl_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        # Row alternating colours
        self.tree.tag_configure("even", background=CARD)
        self.tree.tag_configure("odd",  background=ROW_ALT)

    # ── Upload card ───────────────────────────────────────────────────────────
    def _upload_card(self, parent, icon, title, subtitle, tags, command, side):
        card = tk.Frame(parent, bg=CARD,
                         highlightbackground=ACCENT, highlightthickness=1)
        card.pack(side=side, fill="both", expand=True)

        inner = tk.Frame(card, bg=CARD, pady=14, padx=14)
        inner.pack(fill="both", expand=True)

        # Top row
        top = tk.Frame(inner, bg=CARD)
        top.pack(fill="x")

        icon_box = tk.Frame(top, bg="#EEF1FB", width=38, height=38)
        icon_box.pack(side="left")
        icon_box.pack_propagate(False)
        tk.Label(icon_box, text=icon, font=("Segoe UI", 16),
                 bg="#EEF1FB").place(relx=0.5, rely=0.5, anchor="center")

        txt = tk.Frame(top, bg=CARD)
        txt.pack(side="left", padx=(10, 0))
        tk.Label(txt, text=title, font=FONT_HEAD, fg=TEXT, bg=CARD).pack(anchor="w")
        tk.Label(txt, text=subtitle, font=FONT_SMALL, fg=SUBTEXT, bg=CARD).pack(anchor="w")

        # Tags row
        tag_row = tk.Frame(top, bg=CARD)
        tag_row.pack(side="right")
        for t in tags:
            lbl = tk.Label(tag_row, text=t, font=FONT_SMALL, fg=TAG_FG,
                            bg=TAG_BG, padx=5, pady=2)
            lbl.pack(side="left", padx=2)

        # Divider
        tk.Frame(inner, bg=BORDER, height=1).pack(fill="x", pady=(10, 8))

        # Browse button row
        browse_row = tk.Frame(inner, bg=CARD)
        browse_row.pack(fill="x")

        browse_btn = tk.Button(browse_row, text="＋  Browse Files",
                                font=FONT_SMALL, fg=ACCENT, bg=TAG_BG,
                                activebackground="#D6DCFB", bd=0,
                                relief="flat", cursor="hand2", padx=8, pady=4,
                                command=command)
        browse_btn.pack(side="left")

        # File list area — fixed height, scrollable, wraps files vertically
        list_outer = tk.Frame(inner, bg=CARD, height=90)
        list_outer.pack(fill="x", pady=(6, 0))
        list_outer.pack_propagate(False)

        canvas = tk.Canvas(list_outer, bg=CARD, bd=0, highlightthickness=0, height=90)
        scrollbar = tk.Scrollbar(list_outer, orient="vertical", command=canvas.yview)
        file_frame = tk.Frame(canvas, bg=CARD)

        file_frame.bind("<Configure>",
            lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))

        canvas.create_window((0, 0), window=file_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        card.file_frame = file_frame   # store ref
        card.file_canvas = canvas      # store canvas for scroll reset

        return card

    # ── Browse handlers ───────────────────────────────────────────────────────
    def _browse_resumes(self):
        paths = filedialog.askopenfilenames(
            title="Select Resume PDFs",
            filetypes=[("PDF Files", "*.pdf")]
        )
        for p in paths:
            with open(p, "rb") as f:
                data = f.read()
            name = os.path.basename(p)
            if not any(n == name for n, _ in self.resume_files):
                self.resume_files.append((name, data))
        self._refresh_file_tags(self.resume_card, self.resume_files, self._remove_resume)

    def _browse_aadhaar(self):
        paths = filedialog.askopenfilenames(
            title="Select Aadhaar Files",
            filetypes=[("Images/PDF", "*.pdf *.jpg *.jpeg *.png")]
        )
        for p in paths:
            with open(p, "rb") as f:
                data = f.read()
            name = os.path.basename(p)
            if not any(n == name for n, _ in self.aadhaar_files):
                self.aadhaar_files.append((name, data))
        self._refresh_file_tags(self.aadhaar_card, self.aadhaar_files, self._remove_aadhaar)

    def _remove_resume(self, name):
        self.resume_files = [(n, d) for n, d in self.resume_files if n != name]
        self._refresh_file_tags(self.resume_card, self.resume_files, self._remove_resume)

    def _remove_aadhaar(self, name):
        self.aadhaar_files = [(n, d) for n, d in self.aadhaar_files if n != name]
        self._refresh_file_tags(self.aadhaar_card, self.aadhaar_files, self._remove_aadhaar)

    def _refresh_file_tags(self, card, file_list, remove_fn):
        frame = card.file_frame
        for w in frame.winfo_children():
            w.destroy()
        for name, _ in file_list:
            tag = tk.Frame(frame, bg=TAG_BG, bd=0)
            tag.pack(fill="x", padx=2, pady=2)           # ← vertical, full width
            short = name if len(name) <= 40 else name[:37] + "…"
            tk.Label(tag, text="📄 " + short, font=FONT_SMALL,
                     fg=TEXT, bg=TAG_BG, padx=6, pady=3).pack(side="left")
            x_btn = tk.Button(tag, text="✕", font=("Segoe UI", 8),
                               fg=SUBTEXT, bg=TAG_BG, bd=0, relief="flat",
                               cursor="hand2", padx=3,
                               command=lambda n=name: remove_fn(n))
            x_btn.pack(side="left")

    # ── Extract ───────────────────────────────────────────────────────────────
    def _start_extract(self):
        if not self.resume_files and not self.aadhaar_files:
            messagebox.showwarning("No files", "Please upload at least one file.")
            return
        self.extract_btn.config(state="disabled", text="⏳  Extracting…")
        self.progress_bar.pack(fill="x", pady=(6, 0))
        self.status_lbl.pack(anchor="w")
        self.progress_var.set(0)
        threading.Thread(target=self._extract_worker, daemon=True).start()

    def _extract_worker(self):
        rows = []
        aadhaar_records = []
        total = len(self.resume_files) + len(self.aadhaar_files)
        done  = 0

        # ── Aadhaar ──
        for name, raw in self.aadhaar_files:
            self._set_status(f"Processing Aadhaar: {name}")
            ext = name.lower().rsplit(".", 1)[-1]
            if ext == "pdf":
                primary   = pdf_to_text(raw)
                all_texts = [primary]
            else:
                primary, all_texts = aadhaar_multipass_ocr(raw)
            aadhaar_records.append({
                "aadhaar_name":   get_aadhaar_name(all_texts),
                "aadhaar_mobile": get_aadhaar_phone(primary),
                "address":        get_aadhaar_address(primary),
            })
            done += 1
            self._set_progress(done / total * 100)

        # ── Resumes ──
        def match(rname):
            if not aadhaar_records: return {}
            best, best_score = {}, 0.0
            for rec in aadhaar_records:
                s = name_sim(rname, rec["aadhaar_name"])
                if s > best_score: best_score, best = s, rec
            return best if best_score > 0.55 else {}

        no_aadhaar = not aadhaar_records  # True when no Aadhaar cards uploaded at all

        for name, raw in self.resume_files:
            self._set_status(f"Processing Resume: {name}")
            text    = pdf_to_text(raw)
            rname   = get_resume_name(text)
            matched = match(rname)
            rows.append({
                "Resume Name":    rname,
                "Email":          get_resume_email(text),
                "Resume Mobile":  get_resume_phone(text),
                "Qualification":  get_resume_qualification(text),
                "College":        get_resume_college(text),
                "Aadhaar Name":   matched.get("aadhaar_name", "Not Provided" if no_aadhaar else "Not Matched"),
                "Aadhaar Mobile": matched.get("aadhaar_mobile", ""),
                "Address":        matched.get("address", ""),
            })
            done += 1
            self._set_progress(done / total * 100)

        if not self.resume_files and aadhaar_records:
            for rec in aadhaar_records:
                rows.append({
                    "Resume Name": "", "Email": "", "Resume Mobile": "", "Qualification": "", "College": "",
                    "Aadhaar Name":   rec["aadhaar_name"],
                    "Aadhaar Mobile": rec["aadhaar_mobile"],
                    "Address":        rec["address"],
                })

        self.rows = rows
        self.after(0, self._display_results)

    def _display_results(self):
        # clear table
        for item in self.tree.get_children():
            self.tree.delete(item)

        COLS = ["Resume Name", "Email", "Resume Mobile", "Qualification", "College",
                "Aadhaar Name", "Aadhaar Mobile", "Address"]

        for i, row in enumerate(self.rows):
            vals = tuple(row.get(c, "") or "—" for c in COLS)
            tag  = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", values=vals, tags=(tag,))

        n = len(self.rows)
        self.record_lbl.config(text=f"{n} record{'s' if n != 1 else ''} extracted")
        self.dl_btn.config(state="normal" if self.rows else "disabled")

        self.progress_bar.pack_forget()
        self.status_lbl.pack_forget()
        self.extract_btn.config(state="normal", text="🚀  Extract Data")
        self._set_progress(0)

    # ── Download Excel ────────────────────────────────────────────────────────
    def _download_excel(self):
        if not self.rows:
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="resume_aadhaar_data.xlsx",
            title="Save Excel file"
        )
        if not path:
            return
        data = to_excel_bytes(self.rows)
        if data is None:
            messagebox.showerror("Error", "openpyxl or pandas not installed.")
            return
        with open(path, "wb") as f:
            f.write(data)
        messagebox.showinfo("Saved", f"Excel saved to:\n{path}")

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _set_status(self, msg):
        self.after(0, lambda: self.status_lbl.config(text=msg))

    def _set_progress(self, val):
        self.after(0, lambda: self.progress_var.set(val))

    @staticmethod
    def _bind_hover(widget, normal, hover):
        widget.bind("<Enter>", lambda e: widget.config(bg=hover))
        widget.bind("<Leave>", lambda e: widget.config(bg=normal))


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = DocMatchApp()
    app.mainloop()
